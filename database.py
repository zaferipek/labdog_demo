"""LabDog — Data layer: ORM models, migrations, CRUD operations."""
from __future__ import annotations

import copy
import json
from enum import Enum
from typing import Any, Dict, List, Optional

import streamlit as st

from sqlalchemy import (
    Boolean,
    Date,
    DateTime,
    Enum as SAEnum,
    Float,
    ForeignKey,
    Integer,
    String,
    Text,
    create_engine,
    func,
    or_,
    text,
)
from sqlalchemy.orm import (
    DeclarativeBase,
    Mapped,
    mapped_column,
    relationship,
    sessionmaker,
)


class Base(DeclarativeBase):
    pass


# ``update_material`` için: URL alanlarına dokunulmasın (varsayılan).
_OMIT_DOC_URL = object()


# ══════════════════════════════════════════════════════════════
# ENUMS
# ══════════════════════════════════════════════════════════════

class ExpertiseArea(str, Enum):
    BOYA_FINISH = "Boya/Finish"
    HOT_MELT = "Hot Melt"
    MUREKKEP = "Mürekkep"
    PUD = "PUD"
    PU = "PU"


class ProjectStatus(str, Enum):
    """Proje yaşam döngüsü — 7 aşama (plan: Hammadde tedariği, Lab Test, Validasyon)."""

    FIKIR = "Fikir"
    LITERATURE = "Literatür Taraması"
    HAMMADDE_TEDARIGI = "Hammadde tedariği"
    LAB_TEST = "Lab Test"
    PILOT = "Pilot"
    VALIDASYON = "Validasyon"
    DONE = "Tamamlandı"


_LEGACY_PROJECT_STATUS_STR: dict[str, ProjectStatus] = {
    # Eski DB / Excel / enum adı kalıntıları → yeni .value
    "Laboratuvar Testleri": ProjectStatus.LAB_TEST,
    "LAB_TESTS": ProjectStatus.LAB_TEST,
}


def _coerce_project_status(status_str: str) -> ProjectStatus:
    """İçe aktarma ve güncellemede eski durum metinlerini yeni enum'a çevir."""
    s = (status_str or "").strip()
    if s in _LEGACY_PROJECT_STATUS_STR:
        return _LEGACY_PROJECT_STATUS_STR[s]
    return ProjectStatus(s)


def display_project_status(status_str: str | None) -> str:
    """Liste, path ve filtreler için durumu güncel ProjectStatus.value metnine indirger."""
    s = (status_str or "").strip()
    if s in _LEGACY_PROJECT_STATUS_STR:
        return _LEGACY_PROJECT_STATUS_STR[s].value
    try:
        return ProjectStatus(s).value
    except ValueError:
        return ProjectStatus.FIKIR.value


class ProjectPriority(str, Enum):
    DUSUK = "Düşük"
    ORTA = "Orta"
    YUKSEK = "Yüksek"
    ACIL = "Acil"


class MaterialApprovalStatus(str, Enum):
    ONAYLI = "Onaylı"
    ONAYSIZ = "Onaysız"
    TESTTE = "Testte"
    YOLDA = "Yolda"


class TaskStatus(str, Enum):
    BEKLEMEDE = "Beklemede"
    DEVAM = "Devam Ediyor"
    TAMAMLANDI = "Tamamlandı"
    IPTAL = "İptal"


class TaskPriority(str, Enum):
    DUSUK = "Düşük"
    ORTA = "Orta"
    YUKSEK = "Yüksek"
    ACIL = "Acil"


class UserRole(str, Enum):
    ADMIN = "Admin"
    YONETICI = "Yönetici"
    RND_MUHENDIS = "Ar-Ge Mühendisi"
    RND_UZMANI = "Ar-Ge Uzmanı"
    TEKNIKER = "Tekniker"
    KALITE_KONTROL = "Kalite Kontrol"
    SATIN_ALMA = "Satın Alma"
    GOZLEMCİ = "Gözlemci"


class NoteType(str, Enum):
    DURUM_RAPORU = "Durum Raporu"
    KARAR = "Karar"
    NOT = "Not"
    SORUN = "Sorun"


class FormulationStatus(str, Enum):
    TASLAK = "Taslak"
    TEST = "Test"
    ONAYLANDI = "Onaylandı"
    ARSIV = "Arşiv"


class IngredientRole(str, Enum):
    POLIOL = "Poliol"
    IZOSIYONAT = "İzosiyonat"
    ZINCIR_UZATICI = "Zincir Uzatıcı"
    KATALIZOR = "Katalizör"
    SOLVENT = "Solvent"
    DIGER = "Diğer"


class ProductStatus(str, Enum):
    GELISTIRME = "Geliştirme"
    PILOT_URETIM = "Pilot Üretim"
    SERI_URETIM = "Seri Üretim"
    DURDURULDU = "Durduruldu"


# ══════════════════════════════════════════════════════════════
# ORM MODELS
# ══════════════════════════════════════════════════════════════

class Project(Base):
    __tablename__ = "projects"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    code: Mapped[Optional[str]] = mapped_column(String(20), nullable=True)
    name: Mapped[str] = mapped_column(Text, nullable=False)
    expertise_area: Mapped[ExpertiseArea] = mapped_column(
        SAEnum(ExpertiseArea), nullable=False,
    )
    rd_specialist: Mapped[str] = mapped_column(Text, nullable=False)
    start_date: Mapped[Date] = mapped_column(Date, nullable=False)
    target_date: Mapped[Optional[Date]] = mapped_column(Date, nullable=True)
    status: Mapped[ProjectStatus] = mapped_column(
        SAEnum(ProjectStatus), nullable=False, default=ProjectStatus.FIKIR,
    )
    priority: Mapped[ProjectPriority] = mapped_column(
        SAEnum(ProjectPriority), nullable=False, default=ProjectPriority.ORTA,
    )
    customer: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    description: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    extra_params: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())

    experiments: Mapped[List["Experiment"]] = relationship(
        "Experiment", back_populates="project", cascade="all, delete-orphan",
    )
    tasks: Mapped[List["Task"]] = relationship("Task", back_populates="project")
    notes: Mapped[List["ProjectNote"]] = relationship(
        "ProjectNote", back_populates="project", cascade="all, delete-orphan",
    )
    formulations: Mapped[List["Formulation"]] = relationship(
        "Formulation", back_populates="project", cascade="all, delete-orphan",
    )
    products: Mapped[List["Product"]] = relationship(
        "Product", back_populates="project",
    )


class RawMaterial(Base):
    __tablename__ = "raw_materials"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    code: Mapped[Optional[str]] = mapped_column(String(20), nullable=True)
    name: Mapped[str] = mapped_column(Text, nullable=False)
    stock_name: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    supplier: Mapped[str] = mapped_column(Text, nullable=False)
    brand: Mapped[Optional[str]] = mapped_column(Text)
    function: Mapped[str] = mapped_column("function", Text, nullable=False)
    cas_number: Mapped[Optional[str]] = mapped_column(Text)
    stock_quantity: Mapped[float] = mapped_column(Float, default=0.0)
    unit: Mapped[str] = mapped_column(Text, default="kg")
    arrival_date: Mapped[Optional[Date]] = mapped_column(Date)
    approval_status: Mapped[MaterialApprovalStatus] = mapped_column(
        SAEnum(MaterialApprovalStatus), nullable=False,
        default=MaterialApprovalStatus.ONAYSIZ,
    )
    equivalent_weight: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    nco_content: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    oh_number: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    solid_content: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    safety_data_sheet_url: Mapped[Optional[str]] = mapped_column(Text)
    tds_url: Mapped[Optional[str]] = mapped_column(Text)
    msds_file_path: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    tds_file_path: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    notes: Mapped[Optional[str]] = mapped_column(Text)
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())

    equivalents_source: Mapped[List["MaterialEquivalent"]] = relationship(
        "MaterialEquivalent",
        foreign_keys="MaterialEquivalent.material_id",
        back_populates="material",
    )
    equivalents_target: Mapped[List["MaterialEquivalent"]] = relationship(
        "MaterialEquivalent",
        foreign_keys="MaterialEquivalent.equivalent_id",
        back_populates="equivalent",
    )


class MaterialEquivalent(Base):
    __tablename__ = "material_equivalents"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    material_id: Mapped[int] = mapped_column(
        ForeignKey("raw_materials.id"), nullable=False,
    )
    equivalent_id: Mapped[int] = mapped_column(
        ForeignKey("raw_materials.id"), nullable=False,
    )
    notes: Mapped[Optional[str]] = mapped_column(Text)

    material: Mapped[RawMaterial] = relationship(
        "RawMaterial", foreign_keys=[material_id], back_populates="equivalents_source",
    )
    equivalent: Mapped[RawMaterial] = relationship(
        "RawMaterial", foreign_keys=[equivalent_id], back_populates="equivalents_target",
    )


class Experiment(Base):
    __tablename__ = "experiments"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    code: Mapped[Optional[str]] = mapped_column(String(20), nullable=True)
    project_id: Mapped[int] = mapped_column(
        ForeignKey("projects.id"), nullable=False,
    )
    title: Mapped[str] = mapped_column(Text, nullable=False)
    notes: Mapped[str] = mapped_column(Text, nullable=False)
    date: Mapped[Date] = mapped_column(Date, nullable=False)
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())

    project: Mapped[Project] = relationship("Project", back_populates="experiments")
    test_results: Mapped[List["TestResult"]] = relationship(
        "TestResult", back_populates="experiment", cascade="all, delete-orphan",
    )


class TestResult(Base):
    __tablename__ = "test_results"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    experiment_id: Mapped[int] = mapped_column(
        ForeignKey("experiments.id"), nullable=False,
    )
    test_name: Mapped[str] = mapped_column(Text, nullable=False)
    measured_value: Mapped[float] = mapped_column(Float, nullable=False)
    unit: Mapped[str] = mapped_column(Text, nullable=False)
    observation: Mapped[Optional[str]] = mapped_column(Text)
    is_successful: Mapped[bool] = mapped_column(Boolean, default=True, nullable=False)
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())

    experiment: Mapped[Experiment] = relationship(
        "Experiment", back_populates="test_results",
    )


class Task(Base):
    __tablename__ = "tasks"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    code: Mapped[Optional[str]] = mapped_column(String(20), nullable=True)
    title: Mapped[str] = mapped_column(Text, nullable=False)
    description: Mapped[Optional[str]] = mapped_column(Text)
    project_id: Mapped[Optional[int]] = mapped_column(ForeignKey("projects.id"))
    assignee: Mapped[Optional[str]] = mapped_column(Text)
    status: Mapped[TaskStatus] = mapped_column(
        SAEnum(TaskStatus), nullable=False, default=TaskStatus.BEKLEMEDE,
    )
    priority: Mapped[TaskPriority] = mapped_column(
        SAEnum(TaskPriority), nullable=False, default=TaskPriority.ORTA,
    )
    due_date: Mapped[Optional[Date]] = mapped_column(Date)
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())

    project: Mapped[Optional[Project]] = relationship("Project", back_populates="tasks")


class User(Base):
    __tablename__ = "users"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    code: Mapped[Optional[str]] = mapped_column(String(20), nullable=True)
    username: Mapped[str] = mapped_column(String(255), unique=True, nullable=False)
    password: Mapped[str] = mapped_column(String(255), nullable=False)
    name: Mapped[str] = mapped_column(String(255), nullable=False)
    expertise_group: Mapped[ExpertiseArea] = mapped_column(
        SAEnum(ExpertiseArea), nullable=False,
    )
    role: Mapped[UserRole] = mapped_column(
        SAEnum(UserRole), nullable=False, default=UserRole.GOZLEMCİ,
    )
    is_active: Mapped[bool] = mapped_column(
        Boolean, default=True, nullable=False, server_default="1",
    )
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())


class Setting(Base):
    __tablename__ = "settings"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    key: Mapped[str] = mapped_column(String(255), unique=True, nullable=False)
    value: Mapped[str] = mapped_column(Text, nullable=False)
    updated_at: Mapped[DateTime] = mapped_column(
        DateTime, server_default=func.now(), onupdate=func.now(),
    )


# ══════════════════════════════════════════════════════════════
# USER PREFERENCES  (settings.key = user_prefs.<user_id>, JSON)
# ══════════════════════════════════════════════════════════════

def _user_prefs_storage_key(user_id: int) -> str:
    return f"user_prefs.{int(user_id)}"


DEFAULT_USER_PREFERENCES: Dict[str, Any] = {
    "startup": {
        "target": "dashboard",
        "dash_sub": "cards",
        "project_id": None,
    },
    "ui": {"language": "tr"},
}


def get_user_preferences(user_id: int) -> Dict[str, Any]:
    """Kalıcı tercihler + varsayılanlar (birleştirilmiş kopya)."""
    merged = copy.deepcopy(DEFAULT_USER_PREFERENCES)
    session = SessionLocal()
    try:
        row = (
            session.query(Setting)
            .filter(Setting.key == _user_prefs_storage_key(user_id))
            .first()
        )
        if not row or not (row.value or "").strip():
            return merged
        try:
            data = json.loads(row.value)
        except json.JSONDecodeError:
            return merged
        if isinstance(data.get("startup"), dict):
            merged["startup"].update(data["startup"])
        if isinstance(data.get("ui"), dict):
            merged["ui"].update(data["ui"])
        return merged
    finally:
        session.close()


def save_user_preferences(user_id: int, updates: Dict[str, Any]) -> tuple[bool, str]:
    """Mevcut tercihlerle *updates* birleştirilip kaydedilir."""
    current = get_user_preferences(user_id)
    if "startup" in updates and isinstance(updates["startup"], dict):
        current["startup"].update(updates["startup"])
    if "ui" in updates and isinstance(updates["ui"], dict):
        current["ui"].update(updates["ui"])
    session = SessionLocal()
    try:
        key = _user_prefs_storage_key(user_id)
        row = session.query(Setting).filter(Setting.key == key).first()
        payload = json.dumps(current, ensure_ascii=False)
        if row:
            row.value = payload
        else:
            session.add(Setting(key=key, value=payload))
        session.commit()
        return True, "Tercihler kaydedildi."
    except Exception as exc:
        session.rollback()
        return False, f"Kayıt hatası: {exc}"
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# APP-WIDE AI SETTINGS  (settings.key = labdog.ai, JSON)
# ══════════════════════════════════════════════════════════════

_AI_APP_SETTINGS_KEY = "labdog.ai"


def _read_ai_api_key_from_streamlit_secrets() -> str:
    """Streamlit secrets: OPENAI_API_KEY, LABDOG_AI_API_KEY veya AI_API_KEY."""
    try:
        sec = st.secrets
    except Exception:
        return ""
    for name in ("OPENAI_API_KEY", "LABDOG_AI_API_KEY", "AI_API_KEY"):
        try:
            if name in sec:
                raw = sec[name]
                if raw is not None and str(raw).strip():
                    return str(raw).strip()
        except Exception:
            continue
    return ""


def get_ai_settings() -> Dict[str, Any]:
    """
    Uygulama geneli AI yapılandırması (analiz / gelecek entegrasyonlar için).

    Dönüş anahtarları:
      - ``api_key``: kullanılacak anahtar (Secrets öncelikli; yoksa veritabanı).
      - ``master_prompt``: sistem talimatı metni.
      - ``api_key_source``: ``"secrets"`` | ``"database"`` | ``"none"``.
      - ``has_database_api_key``: DB satırında saklı anahtar var mı (Secrets varken bile).
    """
    master_prompt = ""
    db_api_key = ""
    session = SessionLocal()
    try:
        row = (
            session.query(Setting)
            .filter(Setting.key == _AI_APP_SETTINGS_KEY)
            .first()
        )
        if row and (row.value or "").strip():
            try:
                data = json.loads(row.value)
            except json.JSONDecodeError:
                data = {}
            if isinstance(data, dict):
                master_prompt = str(data.get("master_prompt") or "")
                db_api_key = str(data.get("api_key") or "").strip()
    finally:
        session.close()

    sk = _read_ai_api_key_from_streamlit_secrets()
    has_db = bool(db_api_key)
    if sk:
        return {
            "api_key": sk,
            "master_prompt": master_prompt,
            "api_key_source": "secrets",
            "has_database_api_key": has_db,
        }
    if db_api_key:
        return {
            "api_key": db_api_key,
            "master_prompt": master_prompt,
            "api_key_source": "database",
            "has_database_api_key": True,
        }
    return {
        "api_key": "",
        "master_prompt": master_prompt,
        "api_key_source": "none",
        "has_database_api_key": False,
    }


def set_ai_settings(
    master_prompt: str,
    api_key_update: Optional[str] = None,
) -> tuple[bool, str]:
    """
    ``labdog.ai`` ayarını veritabanında günceller.

    ``api_key_update``: ``None`` → DB'deki API anahtarı aynı kalır; ``\"\"`` → silinir;
    dolu metin → yeni anahtar kaydedilir. Streamlit Secrets ile sağlanan anahtar
    bu fonksiyonla değiştirilemez (çağıran kodda UI engeli gerekir).
    """
    session = SessionLocal()
    try:
        row = (
            session.query(Setting)
            .filter(Setting.key == _AI_APP_SETTINGS_KEY)
            .first()
        )
        current_api = ""
        if row and (row.value or "").strip():
            try:
                prev = json.loads(row.value)
            except json.JSONDecodeError:
                prev = {}
            if isinstance(prev, dict):
                current_api = str(prev.get("api_key") or "").strip()

        new_api = current_api
        if api_key_update is not None:
            new_api = api_key_update.strip()

        payload = json.dumps(
            {
                "api_key": new_api,
                "master_prompt": (master_prompt or "").strip(),
            },
            ensure_ascii=False,
        )
        if row:
            row.value = payload
        else:
            session.add(Setting(key=_AI_APP_SETTINGS_KEY, value=payload))
        session.commit()
        return True, "AI ayarları kaydedildi."
    except Exception as exc:
        session.rollback()
        return False, f"Kayıt hatası: {exc}"
    finally:
        session.close()


class ProjectNote(Base):
    __tablename__ = "project_notes"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    project_id: Mapped[int] = mapped_column(
        ForeignKey("projects.id"), nullable=False,
    )
    author: Mapped[str] = mapped_column(Text, nullable=False)
    note_type: Mapped[NoteType] = mapped_column(
        SAEnum(NoteType), nullable=False, default=NoteType.NOT,
    )
    content: Mapped[str] = mapped_column(Text, nullable=False)
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())

    project: Mapped[Project] = relationship("Project", back_populates="notes")


class Formulation(Base):
    __tablename__ = "formulations"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    code: Mapped[Optional[str]] = mapped_column(String(20), nullable=True)
    project_id: Mapped[int] = mapped_column(
        ForeignKey("projects.id"), nullable=False,
    )
    name: Mapped[str] = mapped_column(Text, nullable=False)
    version: Mapped[int] = mapped_column(Integer, default=1)
    nco_index: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    solid_content: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    ph_target: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    viscosity_target: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    status: Mapped[FormulationStatus] = mapped_column(
        SAEnum(FormulationStatus), nullable=False, default=FormulationStatus.TASLAK,
    )
    notes: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())

    project: Mapped[Project] = relationship("Project", back_populates="formulations")
    ingredients: Mapped[List["FormulationIngredient"]] = relationship(
        "FormulationIngredient", back_populates="formulation",
        cascade="all, delete-orphan",
    )
    products: Mapped[List["Product"]] = relationship(
        "Product", back_populates="formulation",
    )


class FormulationIngredient(Base):
    __tablename__ = "formulation_ingredients"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    formulation_id: Mapped[int] = mapped_column(
        ForeignKey("formulations.id"), nullable=False,
    )
    material_id: Mapped[int] = mapped_column(
        ForeignKey("raw_materials.id"), nullable=False,
    )
    role: Mapped[IngredientRole] = mapped_column(
        SAEnum(IngredientRole), nullable=False, default=IngredientRole.DIGER,
    )
    amount_grams: Mapped[float] = mapped_column(Float, nullable=False)
    equivalent_weight: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    nco_content: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    oh_number: Mapped[Optional[float]] = mapped_column(Float, nullable=True)
    order_index: Mapped[int] = mapped_column(Integer, default=0)

    formulation: Mapped[Formulation] = relationship(
        "Formulation", back_populates="ingredients",
    )
    material: Mapped[RawMaterial] = relationship("RawMaterial")


class Product(Base):
    __tablename__ = "products"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    code: Mapped[Optional[str]] = mapped_column(String(20), nullable=True)
    name: Mapped[str] = mapped_column(Text, nullable=False)
    project_id: Mapped[int] = mapped_column(
        ForeignKey("projects.id"), nullable=False,
    )
    formulation_id: Mapped[Optional[int]] = mapped_column(
        ForeignKey("formulations.id"), nullable=True,
    )
    product_type: Mapped[ExpertiseArea] = mapped_column(
        SAEnum(ExpertiseArea), nullable=False,
    )
    status: Mapped[ProductStatus] = mapped_column(
        SAEnum(ProductStatus), nullable=False, default=ProductStatus.GELISTIRME,
    )
    tds_url: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    sds_url: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    notes: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    created_at: Mapped[DateTime] = mapped_column(DateTime, server_default=func.now())

    project: Mapped[Project] = relationship("Project", back_populates="products")
    formulation: Mapped[Optional[Formulation]] = relationship("Formulation")


# ══════════════════════════════════════════════════════════════
# ENGINE & SESSION
# ══════════════════════════════════════════════════════════════

def get_engine(database_url: str = "sqlite:///labdog_v2.db"):
    return create_engine(database_url, echo=False, future=True)


SessionLocal = sessionmaker(bind=get_engine(), autoflush=False, autocommit=False)


# ══════════════════════════════════════════════════════════════
# INIT, MIGRATION & CODE GENERATION
# ══════════════════════════════════════════════════════════════

_MIGRATIONS: list[str] = [
    "ALTER TABLE users ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT 1",
    "ALTER TABLE users ADD COLUMN code VARCHAR(20)",
    "ALTER TABLE projects ADD COLUMN code VARCHAR(20)",
    "ALTER TABLE projects ADD COLUMN priority VARCHAR DEFAULT 'ORTA'",
    "ALTER TABLE projects ADD COLUMN customer TEXT",
    "ALTER TABLE projects ADD COLUMN extra_params TEXT",
    "ALTER TABLE raw_materials ADD COLUMN code VARCHAR(20)",
    "ALTER TABLE raw_materials ADD COLUMN equivalent_weight FLOAT",
    "ALTER TABLE raw_materials ADD COLUMN nco_content FLOAT",
    "ALTER TABLE raw_materials ADD COLUMN oh_number FLOAT",
    "ALTER TABLE raw_materials ADD COLUMN solid_content FLOAT",
    "ALTER TABLE tasks ADD COLUMN code VARCHAR(20)",
    "ALTER TABLE experiments ADD COLUMN code VARCHAR(20)",
    # Fix records from older migration that used display value instead of enum name
    "UPDATE projects SET priority = 'ORTA' WHERE priority = 'Orta'",
    "UPDATE projects SET priority = 'YUKSEK' WHERE priority = 'Yüksek'",
    "UPDATE projects SET priority = 'DUSUK' WHERE priority = 'Düşük'",
    "UPDATE projects SET priority = 'ACIL' WHERE priority = 'Acil'",
    "ALTER TABLE raw_materials ADD COLUMN msds_file_path TEXT",
    "ALTER TABLE raw_materials ADD COLUMN tds_file_path TEXT",
    "ALTER TABLE raw_materials ADD COLUMN stock_name TEXT",
    # ProjectStatus: Laboratuvar Testleri → Lab Test (7 aşama enum geçişi)
    "UPDATE projects SET status = 'Lab Test' WHERE status = 'Laboratuvar Testleri'",
    "UPDATE projects SET status = 'Lab Test' WHERE status = 'LAB_TESTS'",
]


def init_db() -> None:
    """Create all tables and apply lightweight migrations for existing DBs."""
    engine = get_engine()
    Base.metadata.create_all(bind=engine)
    with engine.connect() as conn:
        for stmt in _MIGRATIONS:
            try:
                conn.execute(text(stmt))
                conn.commit()
            except Exception:
                pass
    _backfill_codes()


def _backfill_codes() -> None:
    """Assign readable codes to any records that don't have one yet."""
    session = SessionLocal()
    try:
        _code_map: list[tuple[type, str]] = [
            (Project, "PRJ"),
            (User, "USR"),
            (RawMaterial, "HM"),
            (Task, "TSK"),
            (Experiment, "EXP"),
        ]
        changed = False
        for model, prefix in _code_map:
            for rec in session.query(model).filter(model.code.is_(None)).all():
                rec.code = f"{prefix}-{rec.id:04d}"
                changed = True
        if changed:
            session.commit()
    except Exception:
        session.rollback()
    finally:
        session.close()


def _set_code(record: Any, prefix: str) -> None:
    """Set readable code after flush (id is available)."""
    record.code = f"{prefix}-{record.id:04d}"


# ══════════════════════════════════════════════════════════════
# SEED DATA
# ══════════════════════════════════════════════════════════════

def seed_initial_data() -> None:
    """Seed database with demo data."""
    engine = get_engine()
    Base.metadata.create_all(bind=engine)
    session = SessionLocal()
    try:
        if session.query(User).count() == 0:
            users = [
                User(
                    username="admin", password="admin", name="Admin User",
                    expertise_group=ExpertiseArea.BOYA_FINISH, role=UserRole.ADMIN,
                ),
                User(
                    username="ahmet", password="1234", name="Ahmet Yılmaz",
                    expertise_group=ExpertiseArea.BOYA_FINISH, role=UserRole.YONETICI,
                ),
                User(
                    username="ayse", password="1234", name="Ayşe Demir",
                    expertise_group=ExpertiseArea.MUREKKEP, role=UserRole.RND_UZMANI,
                ),
                User(
                    username="mehmet", password="1234", name="Mehmet Kaya",
                    expertise_group=ExpertiseArea.PUD, role=UserRole.GOZLEMCİ,
                ),
            ]
            session.add_all(users)
            session.flush()
            for u in users:
                _set_code(u, "USR")

        if session.query(Project).count() == 0:
            m1 = RawMaterial(
                name="Solvent A", supplier="Kimya A.Ş.", function="Çözücü",
                notes="Standart solvent",
            )
            m2 = RawMaterial(
                name="Reçine X", supplier="Polimer Ltd.", function="Bağlayıcı",
                notes="Yüksek dayanım",
            )
            m3 = RawMaterial(
                name="Solvent B (Muadil)", supplier="Global Chem", function="Çözücü",
                notes="Solvent A muadili",
            )
            session.add_all([m1, m2, m3])
            session.flush()
            for m in [m1, m2, m3]:
                _set_code(m, "HM")

            session.add(MaterialEquivalent(
                material_id=m1.id, equivalent_id=m3.id, notes="Birebir değişim",
            ))

            from datetime import date as dt_date, timedelta
            today = dt_date.today()

            p1 = Project(
                name="Yeni Nesil Su Bazlı Boya",
                expertise_area=ExpertiseArea.BOYA_FINISH,
                rd_specialist="Ahmet Yılmaz",
                start_date=today,
                target_date=today + timedelta(days=30),
                status=ProjectStatus.LAB_TEST,
                priority=ProjectPriority.YUKSEK,
                description="Çevre dostu yeni formül geliştirme projesi.",
            )
            p2 = Project(
                name="Hızlı Kuruyan Mürekkep",
                expertise_area=ExpertiseArea.MUREKKEP,
                rd_specialist="Ayşe Demir",
                start_date=today,
                status=ProjectStatus.FIKIR,
                priority=ProjectPriority.ORTA,
                description="Ambalaj sektörü için hızlı kuruma özelliği.",
            )
            session.add_all([p1, p2])
            session.flush()
            for p in [p1, p2]:
                _set_code(p, "PRJ")

            exp1 = Experiment(
                project_id=p1.id,
                title="Viskozite Denemeleri #1",
                notes="Farklı kıvamlaştırıcı oranları denendi.",
                date=today,
            )
            session.add(exp1)
            session.flush()
            _set_code(exp1, "EXP")

            session.add(TestResult(
                experiment_id=exp1.id, test_name="Viskozite (25°C)",
                measured_value=1200.0, unit="cPs",
                observation="Beklenen aralıkta.", is_successful=True,
            ))

            session.add(ProjectNote(
                project_id=p1.id, author="Ahmet Yılmaz",
                note_type=NoteType.DURUM_RAPORU,
                content="Proje başlatıldı. İlk viskozite denemeleri yapıldı, sonuçlar umut verici.",
            ))

        session.commit()
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# AUTH HELPERS
# ══════════════════════════════════════════════════════════════

def get_user_by_username(username: str) -> Optional[User]:
    session = SessionLocal()
    try:
        return session.query(User).filter(User.username == username).one_or_none()
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════

@st.cache_data(ttl=600, show_spinner=False)
def get_dashboard_stats() -> List[Dict[str, Any]]:
    """Per-expertise-area project counts and project lists."""
    session = SessionLocal()
    try:
        results: List[Dict[str, Any]] = []
        for area in ExpertiseArea:
            projects = (
                session.query(Project)
                .filter(Project.expertise_area == area)
                .order_by(Project.created_at.desc())
                .all()
            )
            results.append({
                "expertiseArea": area.value,
                "count": len(projects),
                "projects": [
                    {
                        "id": p.id,
                        "code": p.code or f"PRJ-{p.id:04d}",
                        "name": p.name,
                        "status": p.status.value,
                        "rdSpecialist": p.rd_specialist,
                        "createdAt": p.created_at,
                    }
                    for p in projects
                ],
            })
        return results
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# USER MANAGEMENT
# ══════════════════════════════════════════════════════════════

@st.cache_data(ttl=600, show_spinner=False)
def get_all_users() -> List[Dict[str, Any]]:
    """Return all users ordered by name."""
    session = SessionLocal()
    try:
        users = session.query(User).order_by(User.name).all()
        return [
            {
                "id":              u.id,
                "code":            u.code or f"USR-{u.id:04d}",
                "username":        u.username,
                "name":            u.name,
                "role":            u.role.value,
                "expertise_group": u.expertise_group.value,
                "is_active":       getattr(u, "is_active", True),
                "created_at":      u.created_at,
            }
            for u in users
        ]
    finally:
        session.close()


def add_user(
    name: str, username: str, password: str, role: str, expertise_group: str,
) -> tuple[bool, str]:
    """Insert a new user. Returns (success, message)."""
    session = SessionLocal()
    try:
        if session.query(User).filter(User.username == username).first():
            return False, f"'{username}' kullanıcı adı zaten kullanılıyor."
        user = User(
            name=name, username=username, password=password,
            role=UserRole(role), expertise_group=ExpertiseArea(expertise_group),
            is_active=True,
        )
        session.add(user)
        session.flush()
        _set_code(user, "USR")
        session.commit()
        get_all_users.clear()
        return True, f"'{name}' başarıyla eklendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_user(
    user_id: int,
    name: str,
    username: str,
    role: str,
    expertise_group: str,
    is_active: bool,
) -> tuple[bool, str]:
    """
    Update editable user fields including login username.

    Username must be unique across all users except the row being updated.

    **Görünen ad (Ad Soyad)** ve isteğe bağlı olarak **giriş kullanıcı adı**
    değişince, aynı transaction içinde ``Task.assignee``,
    ``ProjectNote.author`` ve ``Project.rd_specialist`` alanlarında *tam eşleşen*
    metinler yeni değere güncellenir (önce görünen ad, sonra kullanıcı adı eşlemesi).

    Çoğu kayıtta atanan kişi görünen ad ile saklanır; eski veride login adı
    kullanılmışsa o satırlar da güncellenir.
    """
    session = SessionLocal()
    try:
        user = session.query(User).filter(User.id == user_id).first()
        if not user:
            return False, "Kullanıcı bulunamadı."
        uname = username.strip()
        if not uname:
            return False, "Kullanıcı adı boş olamaz."
        taken = (
            session.query(User)
            .filter(User.username == uname, User.id != user_id)
            .first()
        )
        if taken:
            return False, f"'{uname}' kullanıcı adı zaten kullanılıyor."

        old_display = (user.name or "").strip()
        new_display = name.strip()
        old_username = (user.username or "").strip()

        user.username        = uname
        user.name            = new_display
        user.role            = UserRole(role)
        user.expertise_group = ExpertiseArea(expertise_group)
        user.is_active       = is_active

        fts_related_text_changed = False

        if old_display and new_display and old_display != new_display:
            session.query(Task).filter(Task.assignee == old_display).update(
                {Task.assignee: new_display}, synchronize_session=False,
            )
            session.query(ProjectNote).filter(ProjectNote.author == old_display).update(
                {ProjectNote.author: new_display}, synchronize_session=False,
            )
            session.query(Project).filter(Project.rd_specialist == old_display).update(
                {Project.rd_specialist: new_display}, synchronize_session=False,
            )
            fts_related_text_changed = True

        if old_username and uname and old_username != uname:
            session.query(Task).filter(Task.assignee == old_username).update(
                {Task.assignee: uname}, synchronize_session=False,
            )
            session.query(ProjectNote).filter(ProjectNote.author == old_username).update(
                {ProjectNote.author: uname}, synchronize_session=False,
            )
            session.query(Project).filter(Project.rd_specialist == old_username).update(
                {Project.rd_specialist: uname}, synchronize_session=False,
            )
            fts_related_text_changed = True

        session.commit()
        get_all_users.clear()
        if fts_related_text_changed:
            try:
                from search_fts import mark_labdog_fts_stale

                mark_labdog_fts_stale()
            except Exception:
                pass
        return True, f"'{new_display}' başarıyla güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def delete_user(user_id: int) -> tuple[bool, str]:
    """Permanently delete a user. Admin-role users are protected."""
    session = SessionLocal()
    try:
        user = session.query(User).filter(User.id == user_id).first()
        if not user:
            return False, "Kullanıcı bulunamadı."
        if user.role == UserRole.ADMIN:
            return False, "Admin rolündeki kullanıcılar silinemez."
        name = user.name
        session.delete(user)
        session.commit()
        get_all_users.clear()
        return True, f"'{name}' kalıcı olarak silindi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_user_status(user_id: int, is_active: bool) -> tuple[bool, str]:
    """Toggle a user's active / inactive state."""
    session = SessionLocal()
    try:
        user = session.query(User).filter(User.id == user_id).first()
        if not user:
            return False, "Kullanıcı bulunamadı."
        user.is_active = is_active
        session.commit()
        get_all_users.clear()
        state = "aktif" if is_active else "pasif"
        return True, f"'{user.name}' {state} yapıldı."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def _gs_snip(text_val: Optional[str], max_len: int = 120) -> str:
    if not text_val:
        return ""
    s = str(text_val).strip().replace("\n", " ")
    if len(s) <= max_len:
        return s
    return s[: max_len - 1] + "…"


def global_search(
    q: str,
    *,
    limit_per_type: int = 12,
    max_total: int = 80,
) -> List[Dict[str, Any]]:
    """
    Çoklu varlık araması — SQLite LIKE '%q%' (indeks yoksa büyük veride yavaş).

    Dönüş: ``type``, ``title``, ``snippet``, ``code`` (opsiyonel), ``url`` dict.
    ``url`` anahtarları: ``page`` (zorunlu); isteğe bağlı ``project_id``, ``material_id``,
    ``task_id`` (Görevler); proje altı için ``project_tab`` (``ozet``/``form``/``deney``/
    ``gorev``/``notlar``), ``formulation_id``, ``product_id``, ``note_id``, ``experiment_id``.
    """
    raw = (q or "").strip()
    if len(raw) < 2:
        return []

    pat = f"%{raw}%"
    out: List[Dict[str, Any]] = []
    session = SessionLocal()
    try:

        def add(row: Dict[str, Any]) -> None:
            if len(out) >= max_total:
                return
            out.append(row)

        # ── Projeler ─────────────────────────────────────────
        for p in (
            session.query(Project)
            .filter(
                or_(
                    Project.name.like(pat),
                    Project.code.like(pat),
                    Project.rd_specialist.like(pat),
                    Project.customer.like(pat),
                    Project.description.like(pat),
                )
            )
            .order_by(Project.created_at.desc())
            .limit(limit_per_type)
        ):
            code = p.code or f"PRJ-{p.id:04d}"
            add(
                {
                    "type":    "project",
                    "title":   p.name,
                    "snippet": _gs_snip(p.description) or p.rd_specialist or p.customer or "",
                    "code":    code,
                    "url":     {"page": "Projeler", "project_id": str(p.id)},
                }
            )

        # ── Görevler ─────────────────────────────────────────
        for t in (
            session.query(Task)
            .filter(
                or_(
                    Task.title.like(pat),
                    Task.description.like(pat),
                    Task.assignee.like(pat),
                    Task.code.like(pat),
                )
            )
            .order_by(Task.created_at.desc())
            .limit(limit_per_type)
        ):
            code = t.code or f"TSK-{t.id:04d}"
            sub = _gs_snip(t.description) or (t.assignee or "")
            add(
                {
                    "type":    "task",
                    "title":   t.title,
                    "snippet": sub,
                    "code":    code,
                    "url":     {"page": "Görevler", "task_id": str(t.id)},
                }
            )

        # ── Hammaddeler ──────────────────────────────────────
        for m in (
            session.query(RawMaterial)
            .filter(
                or_(
                    RawMaterial.name.like(pat),
                    RawMaterial.code.like(pat),
                    RawMaterial.supplier.like(pat),
                    RawMaterial.brand.like(pat),
                    RawMaterial.function.like(pat),
                    RawMaterial.cas_number.like(pat),
                    RawMaterial.notes.like(pat),
                    RawMaterial.stock_name.like(pat),
                )
            )
            .order_by(RawMaterial.created_at.desc())
            .limit(limit_per_type)
        ):
            code = m.code or f"HAM-{m.id:04d}"
            snip = f"{m.supplier} · {m.function}"
            if m.stock_name:
                snip = f"{m.stock_name} · {snip}"
            add(
                {
                    "type":    "material",
                    "title":   m.name,
                    "snippet": snip,
                    "code":    code,
                    "url":     {"page": "Hammaddeler", "material_id": str(m.id)},
                }
            )

        # ── Formülasyonlar ───────────────────────────────────
        for f in (
            session.query(Formulation)
            .filter(
                or_(
                    Formulation.name.like(pat),
                    Formulation.code.like(pat),
                    Formulation.notes.like(pat),
                )
            )
            .order_by(Formulation.created_at.desc())
            .limit(limit_per_type)
        ):
            code = f.code or f"FRM-{f.id:04d}"
            add(
                {
                    "type":    "formulation",
                    "title":   f.name,
                    "snippet": f"v{f.version} · {f.status.value}",
                    "code":    code,
                    "url":     {
                        "page": "Projeler",
                        "project_id": str(f.project_id),
                        "project_tab": "form",
                        "formulation_id": str(f.id),
                    },
                }
            )

        # ── Ürünler ──────────────────────────────────────────
        for pr in (
            session.query(Product)
            .filter(
                or_(
                    Product.name.like(pat),
                    Product.code.like(pat),
                    Product.notes.like(pat),
                )
            )
            .order_by(Product.created_at.desc())
            .limit(limit_per_type)
        ):
            code = pr.code or f"URN-{pr.id:04d}"
            add(
                {
                    "type":    "product",
                    "title":   pr.name,
                    "snippet": f"{pr.product_type.value} · {pr.status.value}",
                    "code":    code,
                    "url":     {
                        "page": "Projeler",
                        "project_id": str(pr.project_id),
                        "project_tab": "form",
                        "product_id": str(pr.id),
                    },
                }
            )

        # ── Proje notları ────────────────────────────────────
        for n in (
            session.query(ProjectNote)
            .filter(
                or_(
                    ProjectNote.content.like(pat),
                    ProjectNote.author.like(pat),
                )
            )
            .order_by(ProjectNote.created_at.desc())
            .limit(limit_per_type)
        ):
            proj = session.get(Project, n.project_id)
            pname = proj.name if proj else f"Proje #{n.project_id}"
            add(
                {
                    "type":    "note",
                    "title":   _gs_snip(n.content, 80) or "(Not)",
                    "snippet": f"{pname} · {n.author} · {n.note_type.value}",
                    "code":    None,
                    "url":     {
                        "page": "Projeler",
                        "project_id": str(n.project_id),
                        "project_tab": "notlar",
                        "note_id": str(n.id),
                    },
                }
            )

        # ── Deneyler ─────────────────────────────────────────
        for e in (
            session.query(Experiment)
            .filter(
                or_(
                    Experiment.title.like(pat),
                    Experiment.notes.like(pat),
                    Experiment.code.like(pat),
                )
            )
            .order_by(Experiment.created_at.desc())
            .limit(limit_per_type)
        ):
            code = e.code or f"DNY-{e.id:04d}"
            add(
                {
                    "type":    "experiment",
                    "title":   e.title,
                    "snippet": _gs_snip(e.notes, 100),
                    "code":    code,
                    "url":     {
                        "page": "Projeler",
                        "project_id": str(e.project_id),
                        "project_tab": "deney",
                        "experiment_id": str(e.id),
                    },
                }
            )

        return out
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# PROJECT MANAGEMENT
# ══════════════════════════════════════════════════════════════

@st.cache_data(ttl=300, show_spinner=False)
def project_exists(project_id: int) -> bool:
    """Verilen id geçerli bir proje mi?"""
    session = SessionLocal()
    try:
        return (
            session.query(Project).filter(Project.id == int(project_id)).first()
            is not None
        )
    except (TypeError, ValueError):
        return False
    finally:
        session.close()


@st.cache_data(ttl=300, show_spinner=False)
def get_all_projects() -> List[Dict[str, Any]]:
    """Return all projects ordered by created_at desc."""
    session = SessionLocal()
    try:
        projects = session.query(Project).order_by(Project.created_at.desc()).all()
        return [
            {
                "id":             p.id,
                "code":           p.code or f"PRJ-{p.id:04d}",
                "name":           p.name,
                "status":         p.status.value,
                "priority":       (p.priority.value if p.priority else "Orta"),
                "expertise_area": p.expertise_area.value,
                "rd_specialist":  p.rd_specialist,
                "customer":       p.customer or "",
                "start_date":     p.start_date,
                "target_date":    p.target_date,
                "description":    p.description or "",
                "created_at":     p.created_at,
            }
            for p in projects
        ]
    finally:
        session.close()


@st.cache_data(ttl=60, show_spinner=False)
def get_project_by_id(project_id: int) -> Optional[Dict[str, Any]]:
    """Return single project with related counts."""
    session = SessionLocal()
    try:
        p = session.query(Project).filter(Project.id == project_id).first()
        if not p:
            return None
        return {
            "id":               p.id,
            "code":             p.code or f"PRJ-{p.id:04d}",
            "name":             p.name,
            "status":           p.status.value,
            "priority":         (p.priority.value if p.priority else "Orta"),
            "expertise_area":   p.expertise_area.value,
            "rd_specialist":    p.rd_specialist,
            "customer":         p.customer or "",
            "start_date":       p.start_date,
            "target_date":      p.target_date,
            "description":      p.description or "",
            "extra_params":     p.extra_params or "",
            "created_at":       p.created_at,
            "experiment_count": len(p.experiments),
            "task_count":       len(p.tasks),
            "formulation_count": len(p.formulations),
            "product_count":    len(p.products),
            "note_count":       len(p.notes),
        }
    finally:
        session.close()


def _clear_project_caches() -> None:
    get_all_projects.clear()
    get_project_by_id.clear()
    get_dashboard_stats.clear()


def add_project(
    name: str,
    expertise_area: str,
    rd_specialist: str,
    start_date,
    target_date=None,
    status: str = "Fikir",
    priority: str = "Orta",
    customer: str = "",
    description: str = "",
    extra_params: str = "",
) -> tuple[bool, str]:
    """Insert a new project. Returns (success, message)."""
    session = SessionLocal()
    try:
        project = Project(
            name=name.strip(),
            expertise_area=ExpertiseArea(expertise_area),
            rd_specialist=rd_specialist,
            start_date=start_date,
            target_date=target_date,
            status=_coerce_project_status(status),
            priority=ProjectPriority(priority),
            customer=customer.strip() if customer else None,
            description=description.strip() if description else None,
            extra_params=extra_params if extra_params else None,
        )
        session.add(project)
        session.flush()
        _set_code(project, "PRJ")
        session.commit()
        _clear_project_caches()
        return True, f"'{name}' projesi başarıyla oluşturuldu. ({project.code})"
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_project(
    project_id: int,
    name: str,
    rd_specialist: str,
    priority: str,
    customer: str,
    description: str,
    target_date=None,
) -> tuple[bool, str]:
    """Update editable fields of a project."""
    session = SessionLocal()
    try:
        p = session.query(Project).filter(Project.id == project_id).first()
        if not p:
            return False, "Proje bulunamadı."
        p.name          = name.strip()
        p.rd_specialist = rd_specialist
        p.priority      = ProjectPriority(priority)
        p.customer      = customer.strip() if customer else None
        p.description   = description.strip() if description else None
        p.target_date   = target_date
        session.commit()
        _clear_project_caches()
        return True, f"'{name}' başarıyla güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_project_status(
    project_id: int, new_status: str,
) -> tuple[bool, str]:
    """Change only the status of a project. ``new_status`` is the Turkish display string (.value)."""
    try:
        status_enum = _coerce_project_status(new_status)
    except ValueError:
        valid = ", ".join(s.value for s in ProjectStatus)
        return (
            False,
            f"Geçersiz durum: {str(new_status).strip()!r}. Geçerli değerler: {valid}.",
        )

    session = SessionLocal()
    try:
        p = session.query(Project).filter(Project.id == project_id).first()
        if not p:
            return False, "Proje bulunamadı."
        p.status = status_enum
        session.commit()
        _clear_project_caches()
        return True, f"Durum '{status_enum.value}' olarak güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Durum kaydedilemedi: {exc}"
    finally:
        session.close()


def delete_project(project_id: int) -> tuple[bool, str]:
    """Permanently delete a project and all related records."""
    session = SessionLocal()
    try:
        p = session.query(Project).filter(Project.id == project_id).first()
        if not p:
            return False, "Proje bulunamadı."
        name = p.name
        session.delete(p)
        session.commit()
        _clear_project_caches()
        get_project_notes.clear()
        return True, f"'{name}' projesi kalıcı olarak silindi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# PROJECT NOTES
# ══════════════════════════════════════════════════════════════

@st.cache_data(ttl=60, show_spinner=False)
def get_project_notes(project_id: int) -> List[Dict[str, Any]]:
    """Return notes for a project, newest first."""
    session = SessionLocal()
    try:
        notes = (
            session.query(ProjectNote)
            .filter(ProjectNote.project_id == project_id)
            .order_by(ProjectNote.created_at.desc())
            .all()
        )
        return [
            {
                "id":         n.id,
                "author":     n.author,
                "note_type":  n.note_type.value,
                "content":    n.content,
                "created_at": n.created_at,
            }
            for n in notes
        ]
    finally:
        session.close()


def add_project_note(
    project_id: int, author: str, note_type: str, content: str,
) -> tuple[bool, str]:
    """Add a note/status report to a project."""
    session = SessionLocal()
    try:
        note = ProjectNote(
            project_id=project_id,
            author=author,
            note_type=NoteType(note_type),
            content=content.strip(),
        )
        session.add(note)
        session.commit()
        get_project_notes.clear()
        return True, "Not başarıyla eklendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# NCO / OH CALCULATION ENGINE  (pure-Python, no DB required)
# ══════════════════════════════════════════════════════════════

def calculate_nco_index(ingredients: List[Dict[str, Any]]) -> Optional[float]:
    """
    NCO Index = Σ(NCO equivalents) / Σ(OH equivalents)

    Per ingredient:
      NCO_eq = amount_grams / equivalent_weight        (when equivalent_weight is set)
             = amount_grams * nco_content_pct / 4200   (NCO group MW = 42)
      OH_eq  = amount_grams / equivalent_weight
             = amount_grams * oh_number / 56100         (KOH MW = 56.1, ×1000 for g→mg)

    Only roles IZOSIYONAT contribute to NCO_eq.
    Roles POLIOL + ZINCIR_UZATICI contribute to OH_eq.
    Returns None if either sum is zero (division by zero guard).
    """
    nco_eq = 0.0
    oh_eq  = 0.0
    for ing in ingredients:
        amt = float(ing.get("amount_grams") or 0)
        if amt <= 0:
            continue
        role = ing.get("role", "")
        eq_w = ing.get("equivalent_weight") or None
        if role == IngredientRole.IZOSIYONAT.value:
            if eq_w:
                nco_eq += amt / float(eq_w)
            elif ing.get("nco_content"):
                nco_eq += amt * float(ing["nco_content"]) / 4200.0
        elif role in (IngredientRole.POLIOL.value, IngredientRole.ZINCIR_UZATICI.value):
            if eq_w:
                oh_eq += amt / float(eq_w)
            elif ing.get("oh_number"):
                oh_eq += amt * float(ing["oh_number"]) / 56100.0
    if oh_eq == 0:
        return None
    return round(nco_eq / oh_eq, 3)


def calculate_total_weight(ingredients: List[Dict[str, Any]]) -> float:
    """Sum of all ingredient amounts in grams."""
    return sum(float(i.get("amount_grams") or 0) for i in ingredients)


def calculate_weight_fractions(
    ingredients: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Return ingredients list enriched with 'weight_fraction' (0‒100 %).
    Does not mutate the originals.
    """
    total = calculate_total_weight(ingredients)
    result = []
    for ing in ingredients:
        amt = float(ing.get("amount_grams") or 0)
        frac = round(amt / total * 100, 2) if total > 0 else 0.0
        result.append({**ing, "weight_fraction": frac})
    return result


# ══════════════════════════════════════════════════════════════
# FORMULATION MANAGEMENT
# ══════════════════════════════════════════════════════════════

def _clear_formulation_caches() -> None:
    get_all_formulations.clear()
    get_formulation_by_id.clear()
    get_formulation_ingredients.clear()
    get_project_by_id.clear()
    get_dashboard_stats.clear()


@st.cache_data(ttl=120, show_spinner=False)
def get_all_formulations(project_id: int) -> List[Dict[str, Any]]:
    """Return all formulations for a project, newest first."""
    session = SessionLocal()
    try:
        rows = (
            session.query(Formulation)
            .filter(Formulation.project_id == project_id)
            .order_by(Formulation.created_at.desc())
            .all()
        )
        return [
            {
                "id":               f.id,
                "code":             f.code or f"FRM-{f.id:04d}",
                "name":             f.name,
                "version":          f.version,
                "status":           f.status.value,
                "nco_index":        f.nco_index,
                "solid_content":    f.solid_content,
                "ph_target":        f.ph_target,
                "viscosity_target": f.viscosity_target,
                "notes":            f.notes or "",
                "ingredient_count": len(f.ingredients),
                "created_at":       f.created_at,
            }
            for f in rows
        ]
    finally:
        session.close()


@st.cache_data(ttl=60, show_spinner=False)
def get_formulation_by_id(formulation_id: int) -> Optional[Dict[str, Any]]:
    """Return a single formulation with full detail."""
    session = SessionLocal()
    try:
        f = session.query(Formulation).filter(Formulation.id == formulation_id).first()
        if not f:
            return None
        return {
            "id":               f.id,
            "code":             f.code or f"FRM-{f.id:04d}",
            "project_id":       f.project_id,
            "name":             f.name,
            "version":          f.version,
            "status":           f.status.value,
            "nco_index":        f.nco_index,
            "solid_content":    f.solid_content,
            "ph_target":        f.ph_target,
            "viscosity_target": f.viscosity_target,
            "notes":            f.notes or "",
            "ingredient_count": len(f.ingredients),
            "created_at":       f.created_at,
        }
    finally:
        session.close()


@st.cache_data(ttl=60, show_spinner=False)
def get_formulation_ingredients(formulation_id: int) -> List[Dict[str, Any]]:
    """Return all ingredients for a formulation, ordered by order_index."""
    session = SessionLocal()
    try:
        rows = (
            session.query(FormulationIngredient)
            .filter(FormulationIngredient.formulation_id == formulation_id)
            .order_by(FormulationIngredient.order_index)
            .all()
        )
        return [
            {
                "id":               ing.id,
                "formulation_id":   ing.formulation_id,
                "material_id":      ing.material_id,
                "material_name":    ing.material.name if ing.material else "—",
                "material_code":    (ing.material.code or f"HM-{ing.material_id:04d}")
                                    if ing.material else "—",
                "role":             ing.role.value,
                "amount_grams":     ing.amount_grams,
                "equivalent_weight": ing.equivalent_weight,
                "nco_content":      ing.nco_content,
                "oh_number":        ing.oh_number,
                "order_index":      ing.order_index,
            }
            for ing in rows
        ]
    finally:
        session.close()


def add_formulation(
    project_id: int,
    name: str,
    status: str = "Taslak",
    notes: str = "",
    nco_index: Optional[float] = None,
    solid_content: Optional[float] = None,
    ph_target: Optional[float] = None,
    viscosity_target: Optional[float] = None,
) -> tuple[bool, str]:
    """Insert a new formulation. Version is set to max(existing)+1 for the project."""
    session = SessionLocal()
    try:
        existing = (
            session.query(Formulation)
            .filter(Formulation.project_id == project_id)
            .count()
        )
        form = Formulation(
            project_id=project_id,
            name=name.strip(),
            version=existing + 1,
            status=FormulationStatus(status),
            notes=notes.strip() if notes else None,
            nco_index=nco_index,
            solid_content=solid_content,
            ph_target=ph_target,
            viscosity_target=viscosity_target,
        )
        session.add(form)
        session.flush()
        _set_code(form, "FRM")
        session.commit()
        _clear_formulation_caches()
        return True, f"'{name}' formülasyonu oluşturuldu. ({form.code})"
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_formulation(
    formulation_id: int,
    name: str,
    status: str,
    notes: str = "",
    nco_index: Optional[float] = None,
    solid_content: Optional[float] = None,
    ph_target: Optional[float] = None,
    viscosity_target: Optional[float] = None,
) -> tuple[bool, str]:
    """Update editable fields of a formulation."""
    session = SessionLocal()
    try:
        f = session.query(Formulation).filter(Formulation.id == formulation_id).first()
        if not f:
            return False, "Formülasyon bulunamadı."
        f.name             = name.strip()
        f.status           = FormulationStatus(status)
        f.notes            = notes.strip() if notes else None
        f.nco_index        = nco_index
        f.solid_content    = solid_content
        f.ph_target        = ph_target
        f.viscosity_target = viscosity_target
        session.commit()
        _clear_formulation_caches()
        return True, f"'{name}' başarıyla güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_formulation_status(
    formulation_id: int, new_status: str,
) -> tuple[bool, str]:
    """Change only the status of a formulation."""
    session = SessionLocal()
    try:
        f = session.query(Formulation).filter(Formulation.id == formulation_id).first()
        if not f:
            return False, "Formülasyon bulunamadı."
        f.status = FormulationStatus(new_status)
        session.commit()
        _clear_formulation_caches()
        return True, f"Durum '{new_status}' olarak güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def delete_formulation(formulation_id: int) -> tuple[bool, str]:
    """
    Permanently delete a formulation and all its ingredients.
    Blocked if the formulation is linked to a Product.
    """
    session = SessionLocal()
    try:
        f = session.query(Formulation).filter(Formulation.id == formulation_id).first()
        if not f:
            return False, "Formülasyon bulunamadı."
        if f.products:
            return False, (
                "Bu formülasyona bağlı ürün(ler) var; önce ürünleri silin."
            )
        name = f.name
        session.delete(f)
        session.commit()
        _clear_formulation_caches()
        return True, f"'{name}' formülasyonu kalıcı olarak silindi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


# ── FormulationIngredient ──────────────────────────────────────

def add_ingredient(
    formulation_id: int,
    material_id: int,
    role: str,
    amount_grams: float,
    equivalent_weight: Optional[float] = None,
    nco_content: Optional[float] = None,
    oh_number: Optional[float] = None,
) -> tuple[bool, str]:
    """Append an ingredient to a formulation (order_index = last + 1)."""
    session = SessionLocal()
    try:
        last = (
            session.query(func.max(FormulationIngredient.order_index))
            .filter(FormulationIngredient.formulation_id == formulation_id)
            .scalar()
        )
        ing = FormulationIngredient(
            formulation_id=formulation_id,
            material_id=material_id,
            role=IngredientRole(role),
            amount_grams=amount_grams,
            equivalent_weight=equivalent_weight,
            nco_content=nco_content,
            oh_number=oh_number,
            order_index=(last or 0) + 1,
        )
        session.add(ing)
        session.commit()
        get_formulation_ingredients.clear()
        return True, "Bileşen eklendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_ingredient(
    ingredient_id: int,
    role: str,
    amount_grams: float,
    equivalent_weight: Optional[float] = None,
    nco_content: Optional[float] = None,
    oh_number: Optional[float] = None,
) -> tuple[bool, str]:
    """Update mutable fields of a formulation ingredient."""
    session = SessionLocal()
    try:
        ing = (
            session.query(FormulationIngredient)
            .filter(FormulationIngredient.id == ingredient_id)
            .first()
        )
        if not ing:
            return False, "Bileşen bulunamadı."
        ing.role             = IngredientRole(role)
        ing.amount_grams     = amount_grams
        ing.equivalent_weight = equivalent_weight
        ing.nco_content      = nco_content
        ing.oh_number        = oh_number
        session.commit()
        get_formulation_ingredients.clear()
        return True, "Bileşen güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def remove_ingredient(ingredient_id: int) -> tuple[bool, str]:
    """Permanently remove an ingredient from a formulation."""
    session = SessionLocal()
    try:
        ing = (
            session.query(FormulationIngredient)
            .filter(FormulationIngredient.id == ingredient_id)
            .first()
        )
        if not ing:
            return False, "Bileşen bulunamadı."
        session.delete(ing)
        session.commit()
        get_formulation_ingredients.clear()
        return True, "Bileşen silindi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# TASK MANAGEMENT
# ══════════════════════════════════════════════════════════════

def _clear_task_caches() -> None:
    get_all_tasks.clear()
    get_dashboard_stats.clear()


@st.cache_data(ttl=300, show_spinner=False)
def get_all_tasks(project_id: Optional[int] = None) -> List[Dict[str, Any]]:
    """
    Return tasks ordered by created_at desc.
    If project_id is given, return only tasks belonging to that project.
    """
    session = SessionLocal()
    try:
        q = session.query(Task)
        if project_id is not None:
            q = q.filter(Task.project_id == project_id)
        tasks = q.order_by(Task.created_at.desc()).all()
        return [
            {
                "id":          t.id,
                "code":        t.code or f"TSK-{t.id:04d}",
                "title":       t.title,
                "description": t.description or "",
                "project_id":  t.project_id,
                "project_name": (t.project.name if t.project else ""),
                "assignee":    t.assignee or "",
                "status":      t.status.value,
                "priority":    t.priority.value,
                "due_date":    t.due_date,
                "created_at":  t.created_at,
            }
            for t in tasks
        ]
    finally:
        session.close()


def add_task(
    title: str,
    description: str = "",
    project_id: Optional[int] = None,
    assignee: str = "",
    status: str = "Beklemede",
    priority: str = "Orta",
    due_date=None,
) -> tuple[bool, str]:
    """Insert a new task. Returns (success, message)."""
    session = SessionLocal()
    try:
        task = Task(
            title=title.strip(),
            description=description.strip() if description else None,
            project_id=project_id,
            assignee=assignee.strip() if assignee else None,
            status=TaskStatus(status),
            priority=TaskPriority(priority),
            due_date=due_date,
        )
        session.add(task)
        session.flush()
        _set_code(task, "TSK")
        session.commit()
        _clear_task_caches()
        return True, f"'{title}' görevi oluşturuldu. ({task.code})"
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_task(
    task_id: int,
    title: str,
    description: str = "",
    assignee: str = "",
    priority: str = "Orta",
    due_date=None,
) -> tuple[bool, str]:
    """Update editable fields of a task (project and code are immutable)."""
    session = SessionLocal()
    try:
        t = session.query(Task).filter(Task.id == task_id).first()
        if not t:
            return False, "Görev bulunamadı."
        t.title       = title.strip()
        t.description = description.strip() if description else None
        t.assignee    = assignee.strip() if assignee else None
        t.priority    = TaskPriority(priority)
        t.due_date    = due_date
        session.commit()
        _clear_task_caches()
        return True, f"'{title}' başarıyla güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_task_status(task_id: int, new_status: str) -> tuple[bool, str]:
    """Change only the status of a task."""
    session = SessionLocal()
    try:
        t = session.query(Task).filter(Task.id == task_id).first()
        if not t:
            return False, "Görev bulunamadı."
        t.status = TaskStatus(new_status)
        session.commit()
        _clear_task_caches()
        return True, f"Durum '{new_status}' olarak güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def delete_task(task_id: int) -> tuple[bool, str]:
    """Permanently delete a task."""
    session = SessionLocal()
    try:
        t = session.query(Task).filter(Task.id == task_id).first()
        if not t:
            return False, "Görev bulunamadı."
        name = t.title
        session.delete(t)
        session.commit()
        _clear_task_caches()
        return True, f"'{name}' görevi kalıcı olarak silindi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# RAW MATERIAL MANAGEMENT
# ══════════════════════════════════════════════════════════════

def _clear_material_caches() -> None:
    get_all_materials.clear()
    get_material_by_id.clear()


@st.cache_data(ttl=300, show_spinner=False)
def get_all_materials(
    approval_status: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Return all raw materials ordered by name.
    Optionally filter by approval_status value string.
    """
    session = SessionLocal()
    try:
        q = session.query(RawMaterial)
        if approval_status:
            q = q.filter(
                RawMaterial.approval_status == MaterialApprovalStatus(approval_status)
            )
        materials = q.order_by(RawMaterial.name).all()
        return [
            {
                "id":               m.id,
                "code":             m.code or f"HM-{m.id:04d}",
                "name":             m.name,
                "stock_name":       m.stock_name or "",
                "supplier":         m.supplier,
                "brand":            m.brand or "",
                "function":         m.function,
                "cas_number":       m.cas_number or "",
                "stock_quantity":   m.stock_quantity,
                "unit":             m.unit,
                "arrival_date":     m.arrival_date,
                "approval_status":  m.approval_status.value,
                "equivalent_weight": m.equivalent_weight,
                "nco_content":      m.nco_content,
                "oh_number":        m.oh_number,
                "solid_content":    m.solid_content,
                "safety_data_sheet_url": m.safety_data_sheet_url or "",
                "tds_url":              m.tds_url or "",
                "msds_file_path":       m.msds_file_path or "",
                "tds_file_path":        m.tds_file_path or "",
                "notes":            m.notes or "",
                "equivalent_count": len(m.equivalents_source),
                "created_at":       m.created_at,
            }
            for m in materials
        ]
    finally:
        session.close()


@st.cache_data(ttl=120, show_spinner=False)
def get_material_by_id(material_id: int) -> Optional[Dict[str, Any]]:
    """Return a single raw material with its equivalents list."""
    session = SessionLocal()
    try:
        m = session.query(RawMaterial).filter(RawMaterial.id == material_id).first()
        if not m:
            return None
        equivalents = [
            {
                "id":   eq.id,
                "eq_material_id":   e.equivalent_id if e.material_id == material_id else e.material_id,
                "eq_material_name": (
                    e.equivalent.name if e.material_id == material_id else e.material.name
                ),
                "eq_material_code": (
                    (e.equivalent.code or f"HM-{e.equivalent_id:04d}")
                    if e.material_id == material_id
                    else (e.material.code or f"HM-{e.material_id:04d}")
                ),
                "notes": eq.notes or "",
            }
            for eq in m.equivalents_source
            for e in [eq]  # alias
        ]
        return {
            "id":               m.id,
            "code":             m.code or f"HM-{m.id:04d}",
            "name":             m.name,
            "stock_name":       m.stock_name or "",
            "supplier":         m.supplier,
            "brand":            m.brand or "",
            "function":         m.function,
            "cas_number":       m.cas_number or "",
            "stock_quantity":   m.stock_quantity,
            "unit":             m.unit,
            "arrival_date":     m.arrival_date,
            "approval_status":  m.approval_status.value,
            "equivalent_weight": m.equivalent_weight,
            "nco_content":      m.nco_content,
            "oh_number":        m.oh_number,
            "solid_content":    m.solid_content,
            "safety_data_sheet_url": m.safety_data_sheet_url or "",
            "tds_url":              m.tds_url or "",
            "msds_file_path":       m.msds_file_path or "",
            "tds_file_path":        m.tds_file_path or "",
            "notes":            m.notes or "",
            "equivalents":      equivalents,
        }
    finally:
        session.close()


def add_material(
    name: str,
    supplier: str,
    function: str,
    stock_name: str = "",
    brand: str = "",
    cas_number: str = "",
    stock_quantity: float = 0.0,
    unit: str = "kg",
    arrival_date=None,
    approval_status: str = "Onaysız",
    equivalent_weight: Optional[float] = None,
    nco_content: Optional[float] = None,
    oh_number: Optional[float] = None,
    solid_content: Optional[float] = None,
    notes: str = "",
) -> tuple[bool, str]:
    """Insert a new raw material. Returns (success, message)."""
    session = SessionLocal()
    try:
        mat = RawMaterial(
            name=name.strip(),
            stock_name=stock_name.strip() if stock_name else None,
            supplier=supplier.strip(),
            function=function.strip(),
            brand=brand.strip() if brand else None,
            cas_number=cas_number.strip() if cas_number else None,
            stock_quantity=stock_quantity,
            unit=unit,
            arrival_date=arrival_date,
            approval_status=MaterialApprovalStatus(approval_status),
            equivalent_weight=equivalent_weight,
            nco_content=nco_content,
            oh_number=oh_number,
            solid_content=solid_content,
            notes=notes.strip() if notes else None,
        )
        session.add(mat)
        session.flush()
        _set_code(mat, "HM")
        session.commit()
        _clear_material_caches()
        return True, f"'{name}' hammaddesi eklendi. ({mat.code})"
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_material(
    material_id: int,
    name: str,
    supplier: str,
    function: str,
    stock_name: str = "",
    brand: str = "",
    cas_number: str = "",
    stock_quantity: float = 0.0,
    unit: str = "kg",
    arrival_date=None,
    approval_status: str = "Onaysız",
    equivalent_weight: Optional[float] = None,
    nco_content: Optional[float] = None,
    oh_number: Optional[float] = None,
    solid_content: Optional[float] = None,
    notes: str = "",
    safety_data_sheet_url: Any = _OMIT_DOC_URL,
    tds_url: Any = _OMIT_DOC_URL,
) -> tuple[bool, str]:
    """Update all editable fields of a raw material."""
    session = SessionLocal()
    try:
        m = session.query(RawMaterial).filter(RawMaterial.id == material_id).first()
        if not m:
            return False, "Hammadde bulunamadı."
        m.name              = name.strip()
        m.stock_name        = stock_name.strip() if stock_name else None
        m.supplier          = supplier.strip()
        m.function          = function.strip()
        m.brand             = brand.strip() if brand else None
        m.cas_number        = cas_number.strip() if cas_number else None
        m.stock_quantity    = stock_quantity
        m.unit              = unit
        m.arrival_date      = arrival_date
        m.approval_status   = MaterialApprovalStatus(approval_status)
        m.equivalent_weight = equivalent_weight
        m.nco_content       = nco_content
        m.oh_number         = oh_number
        m.solid_content     = solid_content
        m.notes             = notes.strip() if notes else None
        if safety_data_sheet_url is not _OMIT_DOC_URL:
            su = (safety_data_sheet_url or "").strip()
            m.safety_data_sheet_url = su or None
        if tds_url is not _OMIT_DOC_URL:
            tu = (tds_url or "").strip()
            m.tds_url = tu or None
        session.commit()
        _clear_material_caches()
        return True, f"'{name}' başarıyla güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def assign_material_msds_path(
    material_id: int, relative_path: Optional[str],
) -> tuple[bool, str]:
    """Yerel MSDS göreli yolunu günceller; ``None`` dosyayı ve kaydı temizler."""
    import material_files as mf

    session = SessionLocal()
    old: Optional[str] = None
    try:
        m = session.query(RawMaterial).filter(RawMaterial.id == material_id).first()
        if not m:
            return False, "Hammadde bulunamadı."
        old = m.msds_file_path
        if old == relative_path and relative_path is not None:
            return True, "MSDS dosyası güncellendi."
        m.msds_file_path = relative_path
        session.commit()
        _clear_material_caches()
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()

    if old and old != relative_path:
        mf.delete_relative_pdf(old, material_id)
    return True, "MSDS kaydı güncellendi."


def assign_material_tds_path(
    material_id: int, relative_path: Optional[str],
) -> tuple[bool, str]:
    """Yerel TDS göreli yolunu günceller; ``None`` dosyayı ve kaydı temizler."""
    import material_files as mf

    session = SessionLocal()
    old: Optional[str] = None
    try:
        m = session.query(RawMaterial).filter(RawMaterial.id == material_id).first()
        if not m:
            return False, "Hammadde bulunamadı."
        old = m.tds_file_path
        if old == relative_path and relative_path is not None:
            return True, "TDS dosyası güncellendi."
        m.tds_file_path = relative_path
        session.commit()
        _clear_material_caches()
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()

    if old and old != relative_path:
        mf.delete_relative_pdf(old, material_id)
    return True, "TDS kaydı güncellendi."


def update_material_stock(
    material_id: int, quantity: float, unit: str,
) -> tuple[bool, str]:
    """Quick update for stock quantity + unit only."""
    session = SessionLocal()
    try:
        m = session.query(RawMaterial).filter(RawMaterial.id == material_id).first()
        if not m:
            return False, "Hammadde bulunamadı."
        m.stock_quantity = quantity
        m.unit           = unit
        session.commit()
        _clear_material_caches()
        return True, f"Stok {quantity} {unit} olarak güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def delete_material(material_id: int) -> tuple[bool, str]:
    """
    Permanently delete a raw material.
    Blocked if the material is used in any FormulationIngredient.
    """
    session = SessionLocal()
    try:
        m = session.query(RawMaterial).filter(RawMaterial.id == material_id).first()
        if not m:
            return False, "Hammadde bulunamadı."
        used = (
            session.query(FormulationIngredient)
            .filter(FormulationIngredient.material_id == material_id)
            .count()
        )
        if used > 0:
            return False, (
                f"Bu hammadde {used} formülasyonda kullanılıyor; önce formülasyonları güncelleyin."
            )
        name = m.name
        mid = m.id
        session.delete(m)
        session.commit()
        _clear_material_caches()
        try:
            import material_files as mf

            mf.delete_material_upload_folder(mid)
        except Exception:
            pass
        return True, f"'{name}' kalıcı olarak silindi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


# ── MaterialEquivalent (muadil) yönetimi ──────────────────────

def add_material_equivalent(
    material_id: int, equivalent_id: int, notes: str = "",
) -> tuple[bool, str]:
    """
    Link two materials as equivalents (bidirectional).
    Prevents self-links and duplicate pairs.
    """
    if material_id == equivalent_id:
        return False, "Bir hammadde kendisinin muadili olamaz."
    session = SessionLocal()
    try:
        # Check duplicate in either direction
        exists = (
            session.query(MaterialEquivalent)
            .filter(
                (
                    (MaterialEquivalent.material_id == material_id) &
                    (MaterialEquivalent.equivalent_id == equivalent_id)
                ) | (
                    (MaterialEquivalent.material_id == equivalent_id) &
                    (MaterialEquivalent.equivalent_id == material_id)
                )
            )
            .first()
        )
        if exists:
            return False, "Bu muadil ilişkisi zaten tanımlı."
        session.add(
            MaterialEquivalent(
                material_id=material_id,
                equivalent_id=equivalent_id,
                notes=notes.strip() if notes else None,
            )
        )
        session.commit()
        get_material_by_id.clear()
        get_all_materials.clear()
        return True, "Muadil ilişkisi eklendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def remove_material_equivalent(equivalent_record_id: int) -> tuple[bool, str]:
    """Remove a MaterialEquivalent link by its own primary key."""
    session = SessionLocal()
    try:
        eq = (
            session.query(MaterialEquivalent)
            .filter(MaterialEquivalent.id == equivalent_record_id)
            .first()
        )
        if not eq:
            return False, "Muadil kaydı bulunamadı."
        session.delete(eq)
        session.commit()
        get_material_by_id.clear()
        get_all_materials.clear()
        return True, "Muadil ilişkisi kaldırıldı."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# PRODUCT MANAGEMENT  (Onaylanan formülasyondan ürün çıktısı)
# ══════════════════════════════════════════════════════════════

def _clear_product_caches() -> None:
    get_all_products.clear()
    get_project_by_id.clear()
    get_dashboard_stats.clear()


@st.cache_data(ttl=120, show_spinner=False)
def get_all_products(project_id: int) -> List[Dict[str, Any]]:
    """Return all products for a project, newest first."""
    session = SessionLocal()
    try:
        rows = (
            session.query(Product)
            .filter(Product.project_id == project_id)
            .order_by(Product.created_at.desc())
            .all()
        )
        return [
            {
                "id":             p.id,
                "code":           p.code or f"PRD-{p.id:04d}",
                "name":           p.name,
                "product_type":   p.product_type.value,
                "status":         p.status.value,
                "formulation_id": p.formulation_id,
                "formulation_code": (
                    p.formulation.code or f"FRM-{p.formulation_id:04d}"
                    if p.formulation else "—"
                ),
                "formulation_name": p.formulation.name if p.formulation else "—",
                "tds_url":        p.tds_url or "",
                "sds_url":        p.sds_url or "",
                "notes":          p.notes or "",
                "created_at":     p.created_at,
            }
            for p in rows
        ]
    finally:
        session.close()


def add_product(
    project_id: int,
    name: str,
    product_type: str,
    formulation_id: Optional[int] = None,
    status: str = "Geliştirme",
    tds_url: str = "",
    sds_url: str = "",
    notes: str = "",
) -> tuple[bool, str]:
    """Create a product (optionally linked to an approved formulation)."""
    session = SessionLocal()
    try:
        prod = Product(
            project_id=project_id,
            name=name.strip(),
            product_type=ExpertiseArea(product_type),
            formulation_id=formulation_id,
            status=ProductStatus(status),
            tds_url=tds_url.strip() if tds_url else None,
            sds_url=sds_url.strip() if sds_url else None,
            notes=notes.strip() if notes else None,
        )
        session.add(prod)
        session.flush()
        _set_code(prod, "PRD")
        session.commit()
        _clear_product_caches()
        return True, f"'{name}' ürünü oluşturuldu. ({prod.code})"
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def update_product_status(product_id: int, new_status: str) -> tuple[bool, str]:
    """Change only the status of a product."""
    session = SessionLocal()
    try:
        p = session.query(Product).filter(Product.id == product_id).first()
        if not p:
            return False, "Ürün bulunamadı."
        p.status = ProductStatus(new_status)
        session.commit()
        _clear_product_caches()
        return True, f"Durum '{new_status}' olarak güncellendi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


def delete_product(product_id: int) -> tuple[bool, str]:
    """Permanently delete a product."""
    session = SessionLocal()
    try:
        p = session.query(Product).filter(Product.id == product_id).first()
        if not p:
            return False, "Ürün bulunamadı."
        name = p.name
        session.delete(p)
        session.commit()
        _clear_product_caches()
        return True, f"'{name}' ürünü silindi."
    except Exception as exc:
        session.rollback()
        return False, f"Hata: {exc}"
    finally:
        session.close()


# ══════════════════════════════════════════════════════════════
# EXCEL BULK IMPORT  (Proje listesi içe aktarma)
# ══════════════════════════════════════════════════════════════

_IMPORT_COLUMNS = [
    "Proje Adı",
    "Uzmanlık Alanı",
    "Sorumlu",
    "Başlangıç Tarihi",
    "Hedef Tarih",
    "Durum",
    "Öncelik",
    "Müşteri",
    "Açıklama",
]

_IMPORT_AREA_MAP  = {a.value: a for a in ExpertiseArea}
_IMPORT_STATUS_MAP = {s.value: s for s in ProjectStatus}
for _k, _v in _LEGACY_PROJECT_STATUS_STR.items():
    _IMPORT_STATUS_MAP.setdefault(_k, _v)
_IMPORT_PRIO_MAP  = {p.value: p for p in ProjectPriority}


def generate_import_template() -> bytes:
    """
    Return an xlsx file (bytes) with the import template.
    Uses openpyxl.  One header row + one example row.
    """
    try:
        import openpyxl                                        # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415
        import io                                              # noqa: PLC0415

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projeler"

        header_fill  = PatternFill("solid", fgColor="0176D3")
        header_font  = Font(bold=True, color="FFFFFF")
        example_font = Font(italic=True, color="9CA3AF")

        for col_idx, col_name in enumerate(_IMPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = max(16, len(col_name) + 2)

        # Example row
        example = [
            "AquaCoat Geliştirme",          # Proje Adı
            "PUD",                           # Uzmanlık Alanı
            "Ahmet Yılmaz",                  # Sorumlu
            "2026-01-15",                    # Başlangıç Tarihi
            "2026-06-30",                    # Hedef Tarih
            "Fikir",                         # Durum
            "Orta",                          # Öncelik
            "İnterplast A.Ş.",               # Müşteri
            "Örnek proje açıklaması",        # Açıklama
        ]
        for col_idx, val in enumerate(example, start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = example_font

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        raise RuntimeError(
            "openpyxl kütüphanesi yüklü değil. "
            "Terminalde: pip install openpyxl"
        )


def bulk_import_projects(file_bytes: bytes) -> tuple[int, List[str]]:
    """
    Parse an xlsx/xls/csv file and insert valid project rows.
    Returns (success_count, list_of_error_messages).
    """
    try:
        import pandas as pd   # noqa: PLC0415
        import io             # noqa: PLC0415
    except ImportError:
        return 0, ["pandas kütüphanesi yüklü değil. pip install pandas openpyxl"]

    errors: List[str] = []
    success = 0

    try:
        try:
            df = pd.read_excel(io.BytesIO(file_bytes))
        except Exception:
            df = pd.read_csv(io.BytesIO(file_bytes))
    except Exception as exc:
        return 0, [f"Dosya okunamadı: {exc}"]

    required = ["Proje Adı", "Uzmanlık Alanı", "Sorumlu", "Başlangıç Tarihi"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return 0, [f"Zorunlu sütunlar eksik: {', '.join(missing)}"]

    for row_num, row in df.iterrows():
        lineno = int(row_num) + 2  # 1-indexed, +1 for header
        try:
            name = str(row.get("Proje Adı", "")).strip()
            if not name:
                errors.append(f"Satır {lineno}: Proje Adı boş.")
                continue

            area_raw = str(row.get("Uzmanlık Alanı", "")).strip()
            area = _IMPORT_AREA_MAP.get(area_raw)
            if not area:
                errors.append(
                    f"Satır {lineno}: Geçersiz Uzmanlık Alanı '{area_raw}'. "
                    f"Seçenekler: {', '.join(_IMPORT_AREA_MAP.keys())}"
                )
                continue

            specialist = str(row.get("Sorumlu", "")).strip()
            if not specialist:
                errors.append(f"Satır {lineno}: Sorumlu boş.")
                continue

            start_raw = row.get("Başlangıç Tarihi")
            try:
                start_date = pd.to_datetime(start_raw).date()
            except Exception:
                errors.append(f"Satır {lineno}: Geçersiz Başlangıç Tarihi '{start_raw}'.")
                continue

            target_date = None
            target_raw  = row.get("Hedef Tarih")
            if target_raw and str(target_raw).strip() not in ("", "nan", "NaT"):
                try:
                    target_date = pd.to_datetime(target_raw).date()
                except Exception:
                    errors.append(f"Satır {lineno}: Geçersiz Hedef Tarih '{target_raw}'.")
                    continue

            status_raw = str(row.get("Durum", "Fikir")).strip() or "Fikir"
            if status_raw.lower() in ("nan", "nat", "none"):
                status_raw = "Fikir"
            status_obj = _IMPORT_STATUS_MAP.get(status_raw)
            if status_obj is None:
                valid = ", ".join(s.value for s in ProjectStatus)
                errors.append(
                    f"Satır {lineno}: Geçersiz Durum '{status_raw}'. Geçerli: {valid}. "
                    '(Eski şablonda "Laboratuvar Testleri" de kabul edilir.)'
                )
                continue

            prio_raw = str(row.get("Öncelik", "Orta")).strip()
            prio_obj = _IMPORT_PRIO_MAP.get(prio_raw, ProjectPriority.ORTA)

            customer    = str(row.get("Müşteri",   "")).strip() or None
            description = str(row.get("Açıklama",  "")).strip() or None

            ok, msg = add_project(
                name=name,
                expertise_area=area.value,
                rd_specialist=specialist,
                start_date=start_date,
                target_date=target_date,
                status=status_obj.value,
                priority=prio_obj.value,
                customer=customer or "",
                description=description or "",
            )
            if ok:
                success += 1
            else:
                errors.append(f"Satır {lineno}: {msg}")
        except Exception as exc:
            errors.append(f"Satır {lineno}: Beklenmedik hata — {exc}")

    return success, errors


# ══════════════════════════════════════════════════════════════
# BULK IMPORT — USERS
# ══════════════════════════════════════════════════════════════

_USER_IMPORT_COLUMNS = [
    "Kullanıcı Adı",
    "Ad Soyad",
    "Şifre",
    "Rol",
    "Uzmanlık Grubu",
]

_IMPORT_ROLE_MAP = {r.value: r for r in UserRole}
_IMPORT_AREA_MAP_U = {a.value: a for a in ExpertiseArea}


def generate_users_template() -> bytes:
    """Return xlsx template for bulk user import."""
    try:
        import openpyxl                                            # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415
        import io                                                  # noqa: PLC0415

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kullanıcılar"

        header_fill = PatternFill("solid", fgColor="0A8044")
        header_font = Font(bold=True, color="FFFFFF")
        example_font = Font(italic=True, color="9CA3AF")

        for col_idx, col_name in enumerate(_USER_IMPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = max(18, len(col_name) + 2)

        # Rol seçenekleri açıklama sayfası
        ws_info = wb.create_sheet("Seçenekler")
        ws_info.cell(row=1, column=1, value="Geçerli Roller").font = Font(bold=True)
        for i, r in enumerate(UserRole, start=2):
            ws_info.cell(row=i, column=1, value=r.value)
        ws_info.cell(row=1, column=3, value="Geçerli Uzmanlık Grupları").font = Font(bold=True)
        for i, a in enumerate(ExpertiseArea, start=2):
            ws_info.cell(row=i, column=3, value=a.value)

        # Örnek satır
        example = ["fatma.kaya", "Fatma Kaya", "Sifre1234", "Ar-Ge Uzmanı", "PUD"]
        for col_idx, val in enumerate(example, start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = example_font

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        raise RuntimeError("openpyxl kütüphanesi yüklü değil. Terminalde: pip install openpyxl")


def bulk_import_users(file_bytes: bytes) -> tuple[int, List[str]]:
    """
    Parse xlsx/csv and insert valid user rows.
    Returns (success_count, list_of_error_messages).
    """
    try:
        import pandas as pd  # noqa: PLC0415
        import io            # noqa: PLC0415
    except ImportError:
        return 0, ["pandas kütüphanesi yüklü değil. pip install pandas openpyxl"]

    errors: List[str] = []
    success = 0

    try:
        try:
            df = pd.read_excel(io.BytesIO(file_bytes))
        except Exception:
            df = pd.read_csv(io.BytesIO(file_bytes))
    except Exception as exc:
        return 0, [f"Dosya okunamadı: {exc}"]

    required = ["Kullanıcı Adı", "Ad Soyad", "Şifre", "Rol", "Uzmanlık Grubu"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return 0, [f"Zorunlu sütunlar eksik: {', '.join(missing)}"]

    for row_num, row in df.iterrows():
        lineno = int(row_num) + 2
        try:
            username = str(row.get("Kullanıcı Adı", "")).strip()
            name     = str(row.get("Ad Soyad",      "")).strip()
            password = str(row.get("Şifre",          "")).strip()
            role_raw = str(row.get("Rol",            "")).strip()
            area_raw = str(row.get("Uzmanlık Grubu", "")).strip()

            if not all([username, name, password]):
                errors.append(f"Satır {lineno}: Kullanıcı Adı, Ad Soyad ve Şifre zorunludur.")
                continue

            role_obj = _IMPORT_ROLE_MAP.get(role_raw)
            if not role_obj:
                errors.append(
                    f"Satır {lineno}: Geçersiz Rol '{role_raw}'. "
                    f"Seçenekler: {', '.join(_IMPORT_ROLE_MAP.keys())}"
                )
                continue

            area_obj = _IMPORT_AREA_MAP_U.get(area_raw)
            if not area_obj:
                errors.append(
                    f"Satır {lineno}: Geçersiz Uzmanlık Grubu '{area_raw}'. "
                    f"Seçenekler: {', '.join(_IMPORT_AREA_MAP_U.keys())}"
                )
                continue

            ok, msg = add_user(
                name=name,
                username=username,
                password=password,
                role=role_obj.value,
                expertise_group=area_obj.value,
            )
            if ok:
                success += 1
            else:
                errors.append(f"Satır {lineno}: {msg}")
        except Exception as exc:
            errors.append(f"Satır {lineno}: Beklenmedik hata — {exc}")

    return success, errors


# ══════════════════════════════════════════════════════════════
# BULK IMPORT — RAW MATERIALS
# ══════════════════════════════════════════════════════════════

_MATERIAL_IMPORT_COLUMNS = [
    "Hammadde Adı",
    "Stok Adı",
    "Tedarikçi",
    "Marka",
    "Fonksiyon",
    "CAS No",
    "Stok Miktarı",
    "Birim",
    "Onay Durumu",
    "Eşdeğer Ağırlık",
    "NCO İçeriği (%)",
    "OH Sayısı",
    "Katı Madde (%)",
    "Notlar",
]

_IMPORT_APPROVAL_MAP = {s.value: s for s in MaterialApprovalStatus}


def generate_materials_template() -> bytes:
    """Return xlsx template for bulk raw material import."""
    try:
        import openpyxl                                            # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415
        import io                                                  # noqa: PLC0415

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hammaddeler"

        header_fill = PatternFill("solid", fgColor="7B2D8B")
        header_font = Font(bold=True, color="FFFFFF")
        example_font = Font(italic=True, color="9CA3AF")

        for col_idx, col_name in enumerate(_MATERIAL_IMPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = max(18, len(col_name) + 2)

        # Seçenekler sayfası
        ws_info = wb.create_sheet("Seçenekler")
        ws_info.cell(row=1, column=1, value="Geçerli Onay Durumları").font = Font(bold=True)
        for i, s in enumerate(MaterialApprovalStatus, start=2):
            ws_info.cell(row=i, column=1, value=s.value)

        # Örnek satır
        example = [
            "Desmodur W",         # Hammadde Adı
            "DMD-W-01",           # Stok Adı (firma içi, opsiyonel)
            "Covestro",           # Tedarikçi
            "Desmodur W",         # Marka
            "İzosiyonat",         # Fonksiyon
            "5124-30-1",          # CAS No
            "50",                 # Stok Miktarı
            "kg",                 # Birim
            "Onaylı",             # Onay Durumu
            "131",                # Eşdeğer Ağırlık
            "32.0",               # NCO İçeriği (%)
            "",                   # OH Sayısı
            "100",                # Katı Madde (%)
            "Sikloalifatik diizosiyonat",  # Notlar
        ]
        for col_idx, val in enumerate(example, start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = example_font

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        raise RuntimeError("openpyxl kütüphanesi yüklü değil. Terminalde: pip install openpyxl")


def bulk_import_materials(file_bytes: bytes) -> tuple[int, List[str]]:
    """
    Parse xlsx/csv and insert valid raw material rows.
    Returns (success_count, list_of_error_messages).
    """
    try:
        import pandas as pd  # noqa: PLC0415
        import io            # noqa: PLC0415
    except ImportError:
        return 0, ["pandas kütüphanesi yüklü değil. pip install pandas openpyxl"]

    errors: List[str] = []
    success = 0

    try:
        try:
            df = pd.read_excel(io.BytesIO(file_bytes))
        except Exception:
            df = pd.read_csv(io.BytesIO(file_bytes))
    except Exception as exc:
        return 0, [f"Dosya okunamadı: {exc}"]

    required = ["Hammadde Adı", "Tedarikçi", "Fonksiyon"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return 0, [f"Zorunlu sütunlar eksik: {', '.join(missing)}"]

    session = SessionLocal()
    try:
        for row_num, row in df.iterrows():
            lineno = int(row_num) + 2
            try:
                name     = str(row.get("Hammadde Adı", "")).strip()
                supplier = str(row.get("Tedarikçi",    "")).strip()
                func_val = str(row.get("Fonksiyon",    "")).strip()

                if not all([name, supplier, func_val]):
                    errors.append(f"Satır {lineno}: Hammadde Adı, Tedarikçi ve Fonksiyon zorunludur.")
                    continue

                approval_raw = str(row.get("Onay Durumu", "Onaysız")).strip()
                approval_obj = _IMPORT_APPROVAL_MAP.get(
                    approval_raw, MaterialApprovalStatus.ONAYSIZ,
                )

                def _safe_float(val: Any) -> Optional[float]:
                    try:
                        v = str(val).strip()
                        return float(v) if v and v.lower() != "nan" else None
                    except (ValueError, TypeError):
                        return None

                def _safe_str(val: Any) -> Optional[str]:
                    if val is None:
                        return None
                    try:
                        if pd.isna(val):
                            return None
                    except (TypeError, ValueError):
                        pass
                    v = str(val).strip()
                    return v if v and v.lower() != "nan" else None

                material = RawMaterial(
                    name=name,
                    stock_name=_safe_str(
                        row["Stok Adı"] if "Stok Adı" in df.columns else None,
                    ),
                    supplier=supplier,
                    brand=_safe_str(row.get("Marka")),
                    function=func_val,
                    cas_number=_safe_str(row.get("CAS No")),
                    stock_quantity=_safe_float(row.get("Stok Miktarı")) or 0.0,
                    unit=_safe_str(row.get("Birim")) or "kg",
                    approval_status=approval_obj,
                    equivalent_weight=_safe_float(row.get("Eşdeğer Ağırlık")),
                    nco_content=_safe_float(row.get("NCO İçeriği (%)")),
                    oh_number=_safe_float(row.get("OH Sayısı")),
                    solid_content=_safe_float(row.get("Katı Madde (%)")),
                    notes=_safe_str(row.get("Notlar")),
                )
                session.add(material)
                session.flush()
                _set_code(material, "HM")
                success += 1
            except Exception as exc:
                errors.append(f"Satır {lineno}: Beklenmedik hata — {exc}")

        session.commit()
        if success > 0:
            get_all_projects.clear()  # dashboard stats may reference materials indirectly
    except Exception as exc:
        session.rollback()
        return 0, [f"Veritabanı hatası: {exc}"]
    finally:
        session.close()

    return success, errors


# ══════════════════════════════════════════════════════════════
# FTS5 — arama indeksini veriyle senkron tutma (ORM commit sonrası)
# ══════════════════════════════════════════════════════════════


def _register_labdog_fts_listeners() -> None:
    """Aranabilir modellerde commit olunca FTS tam yenileme kuyruğa alınır."""
    from sqlalchemy import event
    from sqlalchemy.orm import Session as SaSession

    _FTS_TRACKED_MODELS = (
        Project,
        Task,
        RawMaterial,
        Formulation,
        Product,
        ProjectNote,
        Experiment,
    )

    @event.listens_for(SaSession, "after_flush")
    def _labdog_fts_after_flush(session, flush_context) -> None:
        for bucket in (session.new, session.dirty, session.deleted):
            for obj in bucket:
                if isinstance(obj, _FTS_TRACKED_MODELS):
                    session.info["_labdog_fts_pending"] = True
                    return

    @event.listens_for(SaSession, "after_commit")
    def _labdog_fts_after_commit(session) -> None:
        if not session.info.pop("_labdog_fts_pending", False):
            return
        try:
            from search_fts import mark_labdog_fts_stale

            mark_labdog_fts_stale()
        except Exception:
            pass

    @event.listens_for(SaSession, "after_rollback")
    def _labdog_fts_after_rollback(session) -> None:
        session.info.pop("_labdog_fts_pending", None)


_register_labdog_fts_listeners()
